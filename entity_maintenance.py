#!/usr/bin/env python
# -*- coding: utf8 -*-
# 
# © 2013-2014 Institute of Mathematics and Computer Science, University of Latvia
# (LU aģentūra "Latvijas Universitātes Matemātikas un informātikas institūts")
#
# All rights reserved.

from __future__ import unicode_literals

import sys, json, getopt
from collections import defaultdict

sys.path.append("./src")
from db_config import api_conn_info
from SemanticApiPostgres import SemanticApiPostgres, PostgresConnection
from DocumentUpload import inflectEntity, orgAliases
from openpyxl import load_workbook

def fetch_entities():
    sql = "select e.entityid, e.name, e.category, e.dataset, e.nameinflections, array_agg(n.name) aliases, min(i.outerid) ids from entities e \
            left outer join entityothernames n on e.entityid = n.entityid \
            left outer join entityouterids i on e.entityid = i.entityid \
            where e.deleted is false \
            group by e.entityid, e.name, e.category, e.dataset, e.nameinflections\
            order by e.dataset, e.category, e.entityid"
    res = api.api.query(sql, None )

    entities = defaultdict(lambda: defaultdict(list)) # multimaps: dataset->category->list of entities there

    for entity in res:
        entities[entity.dataset][entity.category].append(entity)

    for dataset, dataset_entities in entities.items():
        for category, category_entities in dataset_entities.items():
            with open('./entity_fixtures/%s-%d-%d.json' % (api_conn_info.get('dbname'), dataset, category), 'wt') as file_handle:
                file_handle.write('[\n')
                for entity in category_entities:
                    d = entity.__dict__
                    nameinflections = d.get('nameinflections')
                    if nameinflections:
                        try:
                            nameinflections = json.loads(nameinflections)
                        except Exception as e:
                            nameinflections = None
                    d['nameinflections'] = nameinflections
                    file_handle.write('%s,\n' % json.dumps(d, ensure_ascii=False))
                file_handle.write('{}]')

def create_dates():
    m_names = ['janvāris', 'februāris', 'marts', 'aprīlis', 'maijs', 'jūnijs', 'jūlijs', 'augusts', 'septembris', 'oktobris', 'novembris', 'decembris']
    m_names = ['janvāris', 'februāris', 'marts', 'aprīlis', 'maijs', 'jūnijs', 'jūlijs', 'augusts', 'septembris', 'oktobris', 'novembris', 'decembris']

    # def insertEntity(self, name, othernames, category, outerids=[], inflections = None, hidden = False, commit = True):
    for year in range(1900, 2021):
        name = '%d. gads' % year
        aliases = [name, '%d' % year, '%d.' % year, '%d. g' % year, '%d. g.' % year]        
        inflections = inflectEntity(name, 'date')
        api.insertEntity(name, aliases, 6, inflections = inflections, commit = False)
        print(name, aliases, inflections)

        for month in range(1,13):
            name = '%d. gada %s' % (year, m_names[month-1])
            aliases = [name, '%d.%02d' % (year, month), '%d.%02d.' % (year, month)]
            inflections = inflectEntity(name, 'date')
            if month<10:
                aliases.append('%d.%d' % (year, month))
                aliases.append('%d.%d.' % (year, month))
            api.insertEntity(name, aliases, 6, inflections = inflections, commit = False)
            #print(name, aliases, inflections)

    conn.commit()
    print("Dates loaded!")

def load_education():
    education_xls = "/Users/pet/Dropbox/Resursi/Leksiskie dati/izglītības iestādes14082014.xlsm"
    print('Loading excel...')
    book = load_workbook(education_xls, data_only = True)
    sheet = book.get_sheet_by_name(name = "Nosaukumi un alias")
    print('Excel in memory')
    rows = sheet.range("A2:B2000")   #FIXME - hardkodēts range, kurā tajā excelī ir dati

    entitynames = defaultdict(set) # multimaps: entity canonical name -> set of aliases
    for row in rows:
        representative = row[1].value
        alias = row[0].value
        if representative:
            entitynames[representative].add(representative)
            if alias:
                entitynames[representative].add(alias)
                if 'vidusskola' in alias:
                    entitynames[representative].add(alias.replace('vidusskola', 'vsk.'))
                    entitynames[representative].add(alias.replace('vidusskola', 'vsk'))

    for representative, aliases in entitynames.items():
        inflections = inflectEntity(representative, 'organization')
        api.insertEntity(representative, aliases, 2, inflections = inflections, commit = False)
        # print("%s -> %s" % (representative, aliases))
    conn.commit()
    print("Education institutions loaded!")

# Saģenerē dažādus 'apcirptos' aliasus
def aliasname(name):
    result = set()
    name = name.strip()
    result.add(name)
    if name.startswith('"') and name.endswith('"'):
        result.update(aliasname(name[1:-1]))
    if '"' not in name:
        result.add('"' + name + '"')
    if name.startswith('Partija') or name.startswith('partija'):
        result.update(aliasname(name[8:])) # tas pats tikai bez partija sākumā
    if name.startswith('Politiskā partija'):
        result.update(aliasname(name[10:])) # tas pats tikai bez Politiskā sākumā (partija paliek)
    return result


def load_parties():
    party_xls = "/Users/pet/Dropbox/Resursi/Leksiskie dati/Partiju_lielaako_saraksts.xlsx"
    print('Loading excel...')
    book = load_workbook(party_xls, data_only = True)
    sheet = book.get_sheet_by_name(name = "Sheet1")
    print('Excel in memory')
    rows = sheet.range("A2:D204")   #FIXME - hardkodēts range, kurā tajā excelī ir dati

    entitynames = defaultdict(set) # multimaps: entity canonical name -> set of aliases
    last_representative = None
    for row in rows:
        representative = row[0].value
        alias = row[1].value
        abbr1 = row[2].value
        abbr2 = row[3].value
        if representative:
            last_representative = representative 
        else:
            representative = last_representative # ja A kolonna ir tukša, tad tas nozīmē, ka entītija turpinās

        entitynames[representative].update(aliasname(representative))

        if alias:
            entitynames[representative].update(aliasname(alias))
        if abbr1:
            entitynames[representative].add(abbr1)
        if abbr2:
            entitynames[representative].add(abbr2)

    for representative, aliases in entitynames.items():
        inflections = inflectEntity(representative, 'organization')
        api.insertEntity(representative, aliases, 2, inflections = inflections, commit = False)
        #print("%s %s-> %s" % (representative, inflections, aliases))
    conn.commit()
    print("Parties loaded!")

def load_cv_entities(filename):
    with open(filename) as f:
        cv_entities = json.load(f)

    query = """
    SELECT e.entityid, NAME, dataset, cv_status, i.outerid
    FROM entities e JOIN entityouterids i ON e.entityid = i.entityid
    where i.outerid = %s
    """

    for entity in cv_entities:
        name = entity.get('representative')
        aliases = entity.get('aliases')
        category = 0
        if entity.get('type') == 'organization':
            category = 2
            aliases = set(orgAliases(name)).union(aliases)
        if entity.get('type') == 'person':
            category = 3        
        
        res = api.api.query(query, (entity.get('uqid'),) )
        if res and res[0].dataset == 3:
            continue
        print('Skatamies uz %s - ID %s' % (name,entity.get('uqid')))
        print(res)

        api.insertEntity(name, aliases, category, outerids=[entity.get('uqid')], cv_status=1, source='LETA 20150122 CV apdeita entītiju ielāde', inflections = json.dumps(entity.get('inflections'), ensure_ascii=False), commit = False)
    conn.commit()
    print("Entities from file %s loaded!" % filename)

def load_entities(filename):
    with open(filename) as f:
        entities = json.load(f)

    for counter, entity in enumerate(entities):
        name = entity.get('name')
        aliases = set(entity.get('aliases'))
        category = entity.get('category')
        inflections = json.dumps(entity.get('nameinflections'), ensure_ascii=False)
        outerids=[]
        if entity.get('ids'):
            outerids=[entity.get('ids')]
        api.insertEntity(name, aliases, category, outerids=outerids, inflections = inflections, commit = False)
        # print(name, aliases, category, inflections)
        if counter % 100 == 99:
            print('%s\n' % (counter+1,))
            conn.commit()
    conn.commit()
    print("Entities from file %s loaded!" % filename)

def reinflect_entities():
    print('Re-inflecting all entities!')
    res = api.api.query("select entityid, name, categorynameeng, nameinflections from entities e JOIN entitycategories c ON e.category = c.categoryid", None )
    
    # res = api.api.query("SELECT entityid, NAME, categorynameeng, nameinflections FROM entities e JOIN entitycategories c ON e.category = c.categoryid where entityid = 2585016", None )
    # res = api.api.query("SELECT entityid, NAME, categorynameeng, nameinflections FROM entities e JOIN entitycategories c ON e.category = c.categoryid where category = 1 and dataset = 1", None )    
    # res = api.api.query("SELECT entityid, NAME, categorynameeng, nameinflections FROM entities e JOIN entitycategories c ON e.category = c.categoryid where category = 3", None )    

    for counter, entity in enumerate(res):
        inflections = inflectEntity(entity.name, entity.categorynameeng)
        gender = json.loads(entity.nameinflections).get('Dzimte')
        if gender:
            inflections = json.loads(inflections)
            inflections['Dzimte'] = gender
            inflections = json.dumps(inflections, ensure_ascii=False)
        api.api.insert("update entities set nameinflections = %s where entityid = %s", (inflections, entity.entityid))
        if counter % 1000 == 999:
            print('%s' % (counter+1,))
            conn.commit()

    conn.commit()
    print('Entity re-inflection done!')


def clear_db():
    # if api_conn_info["dbname"] != 'accept_test':
    #     print "Error: clearing allowed only for accept_test"
    #     quit()

    sql = """
-- Izdzēšam visus neblesotos freimus
DELETE FROM summaryframeroledata WHERE frameid IN 
    (SELECT A.frameid FROM summaryframes AS A WHERE A.blessed IS NULL);
DELETE FROM summaryframedata WHERE frameid IN 
    (SELECT A.frameid FROM summaryframes AS A WHERE A.blessed IS NULL);
DELETE FROM summaryframes WHERE blessed IS NULL;

DELETE FROM framedata WHERE frameid IN 
    (SELECT A.frameid FROM frames AS A WHERE A.blessed IS NULL);
DELETE FROM frames WHERE blessed IS NULL;


-- Izdzēšam entītes, uz kurām vairs neviens freims neatsaucas
DELETE FROM entityouterids WHERE entityid IN 
    (SELECT entityid FROM entities
    WHERE NOT entityid IN (SELECT DISTINCT entityid FROM framedata)
    AND NOT entityid IN (SELECT DISTINCT entityid FROM summaryframeroledata)
    AND dataset = 0);

DELETE FROM entityothernames WHERE entityid IN 
    (SELECT entityid FROM entities
    WHERE NOT entityid IN (SELECT DISTINCT entityid FROM framedata)
    AND NOT entityid IN (SELECT DISTINCT entityid FROM summaryframeroledata)
    AND dataset = 0);

DELETE FROM entitymentions WHERE entityid IN 
    (SELECT entityid FROM entities
    WHERE NOT entityid IN (SELECT DISTINCT entityid FROM framedata)
    AND NOT entityid IN (SELECT DISTINCT entityid FROM summaryframeroledata)
    AND dataset = 0);

DELETE FROM entities
WHERE NOT entityid IN (SELECT DISTINCT entityid FROM framedata)
AND NOT entityid IN (SELECT DISTINCT entityid FROM summaryframeroledata)
AND dataset = 0;

DELETE FROM dirtyentities;
    """
    api.api.insert(sql, None, returning=False, commit=True)

    print('Database %s cleared of all non-blessed frames and orphan autogenerated entities!' % (api_conn_info["dbname"],))

def describe_namesakes():
    namesake_sql = """
SELECT e.entityid, NAME, description, i.outerid
FROM entities e
LEFT OUTER JOIN entityouterids i ON e.entityid = i.entityid
WHERE NAME IN (
    SELECT NAME FROM entities
    WHERE NOT deleted AND category = 3 AND dataset = 3
    GROUP BY NAME
    HAVING count(*) > 1
    )
ORDER BY category, NAME, dataset
    """

    entityframe_sql = """
SELECT f.frameid, frametypeid, sentenceid, roleid, NAME, nameinflections FROM 
frames f
JOIN framedata d ON f.frameid = d.frameid
JOIN entities e ON d.entityid = e.entityid
WHERE f.frameid IN 
    (SELECT frameid FROM framedata WHERE entityid = %s)
AND frametypeid IN (7,9,10)
order by sentenceid desc
    """

    entities = api.api.query(namesake_sql, None)
    for counter, entity in enumerate(entities):
        
        frames = api.api.query(entityframe_sql, (entity.entityid,))
        best_descriptor = ''
        for frame in frames:
            if frame.frametypeid == 7 and frame.roleid == 2: # Nodarbošanās
                best_descriptor = frame.name
                break
            if frame.frametypeid in (9,10) and frame.roleid == 3: # Amats, darba sākums
                best_descriptor = frame.name
                if best_descriptor in ('vadītājs', 'direktors', 'rīkotājdirektors', 'izpilddirektors', 'direktora vietnieks', 'valdes loceklis', 'valdes priekšsēdētājs', 'padomes loceklis', 'prezidents', 'priekšsēdētājs', 'priekšsēdētāja', 'priekšnieks', 'priekšniece', 'īstenais loceklis', 'īpašnieks', 'līdzīpašnieks'):
                    for frame_firma in frames:
                        if frame_firma.frameid == frame.frameid and frame_firma.roleid == 2: # šī paša freima ietvaros...
                            best_descriptor += ', ' + frame_firma.name 
                    
                break
        print('%s\t%s\t%s\t%s' % (entity.entityid, entity.outerid, entity.name, best_descriptor))
        if counter % 1000 == 999:
            print('%s' % (counter+1,))
        # if counter > 40:
        #     break


def retokenize_main_name():
    print('Verifying if proper tokenization of primary name exists in aliases')
    sql = """
        select e.entityid, e.name, e.category, e.dataset, e.nameinflections, c.categorynameeng, array_agg(n.name) aliases from entities e 
        left outer join entityothernames n on e.entityid = n.entityid
        join entitycategories c ON e.category = c.categoryid
        where e.deleted is false
        group by e.entityid, e.name, e.category, e.dataset, e.nameinflections, c.categorynameeng
        order by e.dataset, e.category, e.entityid
        """
    res = api.api.query(sql, None )
    print('Query done...')

    names_sql = "INSERT INTO EntityOtherNames(EntityID, Name) VALUES (%s, %s)"

    for counter, entity in enumerate(res):
        inflections = inflectEntity(entity.name, entity.categorynameeng)
        inflections = json.loads(inflections)
        if inflections.get('Nominatīvs') != entity.name and inflections.get('Nominatīvs') not in entity.aliases:
            print('Insertojam aliasu %s entītijai %s' % (inflections.get('Nominatīvs'), entity.name))            
            api.api.insert(names_sql, (entity.entityid, inflections.get('Nominatīvs')) )

        if counter % 1000 == 999:
            print('%s' % (counter+1,))
            conn.commit()

    conn.commit()
    print('Entity re-tokenization done!')


def main():
    # create_dates()
    # load_education()
    # load_parties()
    # load_cv_entities("entity_fixtures/gold/Organizācijas no LETA20150122.json")
    # load_cv_entities("entity_fixtures/gold/Personas no LETA20150122.json")
    # load_entities("entity_fixtures/gold/Vietas no LĢIS.json")
    # load_entities("entity_fixtures/gold/Personas no firmu exceļa.json")
    # load_entities("entity_fixtures/gold/Organizācijas no firmu exceļa.json")
    # fetch_entities()
    # reinflect_entities()
    # describe_namesakes()
    retokenize_main_name()
    print('Done!')

if __name__ == "__main__":
    options, remainder = getopt.getopt(sys.argv[1:], '', ['help', 'cleardb', 'reinflect', 'database='])
    cleardb = False
    reinflect = False
    for opt, arg in options:
        if opt == '--help':
            print('Entity maintenance script')
            print('')
            print('Usage: runs whatever is uncommented in code, handle everything manually!')
            print('--database=<dbname>   overrides the database name from the one set in db_config.py')
            print('--reinflect           reinflect entities')
            print('--cleardb             instead of normal operation, runs a script to clear the selected DB of unblessed frames and entities')
            quit()
        elif opt == '--database':
            api_conn_info["dbname"] = arg
        elif opt == '--cleardb':
            cleardb = True
        elif opt == '--reinflect':
            reinflect = True

    conn = PostgresConnection(api_conn_info, dataset=4)
    api = SemanticApiPostgres(conn)

    if cleardb:
        clear_db()
    else:
        if reinflect:
            reinflect_entities()
    else:
        main()