#!/usr/bin/env python
# -*- coding: utf8 -*-
# 
# © 2013-2014 Institute of Mathematics and Computer Science, University of Latvia
# (LU aģentūra "Latvijas Universitātes Matemātikas un informātikas institūts")
#
# All rights reserved.

from openpyxl import load_workbook

from pprint import pprint
from collections import Counter

class FrameInfo(object):
    def __init__(self, fname):
        self.matrix = FrameMatrix(fname)

        self.frame_types = None
        self.elem_types = None
        self.elem_core = None
        self.elem_other = None

        self.load()

    def load(self):
        self.frame_types = dict(enumerate(self.matrix.frame_types))

        self.elem_types = {}

        for fr_type_id, elem_list in enumerate(self.matrix.frame_elements):
            self.elem_types[fr_type_id] = [item for item in elem_list if item]

        elem_core = {}
        elem_other = {}

        for fr_type_id in self.elem_types:
            t_elem_info = self.matrix.elem_class[fr_type_id]

            elem_core.setdefault(fr_type_id, [])
            elem_other.setdefault(fr_type_id, [])

            for fr_elem_id, fr_elem_class in enumerate(t_elem_info):

                if fr_elem_class:      # ignore empty elements
                    if fr_elem_class == "Core":
                        elem_core[fr_type_id].append(fr_elem_id)
                    else:
                        elem_other[fr_type_id].append(fr_elem_id)

        self.elem_core = elem_core 
        self.elem_other = elem_other

    def type_name_from_id(self, fr_type_id):
        return self.frame_types[fr_type_id]

    def elem_name_from_id(self, fr_type_id, el_type_id):
        try:
            return self.elem_types[fr_type_id][el_type_id]
        except IndexError:
            print('elem_name_from_id bad indexes',fr_type_id, el_type_id)
            return ""

    def get_core_elements(self, fr_type_id):
        return self.elem_core[fr_type_id]

    def get_other_elements(self, fr_type_id):
        return self.elem_other[fr_type_id]

    def all_frame_ids(self):
        return self.frame_types.keys()

    def print_frame_info(self):
        for fr_type_id in self.frame_types:
            print("%s\t%s" % (fr_type_id, self.frame_types[fr_type_id]))

            print("\tAll:\t" + "\t".join(self.elem_types[fr_type_id]))
            print()

            print("\tCore:\t" + "\t".join(str(item) for item in self.get_core_elements(fr_type_id)))
            print("\t\t" + "\t".join(self.elem_name_from_id(fr_type_id, ind) for ind in self.get_core_elements(fr_type_id)))
            print("\tOther:\t" + "\t".join(str(item) for item in self.get_other_elements(fr_type_id)))
            print("\t\t" + "\t".join(self.elem_name_from_id(fr_type_id, ind) for ind in self.get_other_elements(fr_type_id)))

            print()


class FrameMatrix(object):

# TESTING:
#  - that existing file can be loaded
#     - if it is XLSX file
#     - if the matrix sheet exists
#     - if the format is correct

#  - that values in the file are read correctly
#     - only 1 line
#     - 2 lines (first, last)

    def __init__(self, fname):
        self.frame_types = []
        self.frame_elements = []
        self.elem_class = []

        self.load(fname)

    def load(self, fname):
        """
        Loads FrameMatrix from an XLSX file.
        """

        book = load_workbook(fname, data_only = True)
        sheet = book.get_sheet_by_name(name = "FrameMatrix")
      
        # add a check for empty rows in the XLS file (!)

        # FIXME - šo vajag ņemt no eksceļa pa tiešo kautkā
        frame_cnt = 26      # 0..26 = (total of 27 frames)

        rows = sheet.range("B%i:L%i" % (2, 2+frame_cnt))
        for row in rows:
            self.frame_types.append(row[0].value)
            self.frame_elements.append([cell.value for cell in row[1:]])


        # check that prev. row is empty
        matrix2_pos = 30 # FIXME - šo arī jāņemo no exceļa, savādāk salūzt ar citu datukopu

        empty_line2 = sheet.range("B%i:L%i" % (matrix2_pos-1, matrix2_pos-1))[0]
        if not all(cell.value is None for cell in empty_line2):
            raise Exception("Row %i (before the data table) must be empty.")

        # read matrix2
        rows = sheet.range("B%i:L%i" % (matrix2_pos, matrix2_pos+frame_cnt))
        for (pos, row) in enumerate(rows):
            if row[0].value != self.frame_types[pos]:
                raise Exception("Frame name mismatch in reading Matrix #2 - '%s' vs '%s' at row %d" % (row[0].value, self.frame_types[pos], pos))

            self.elem_class.append([cell.value for cell in row[1:]])

    def get_frame_info(self, pos):
        return (self.frame_types[pos], self.frame_elements[pos], self.elem_class[pos], \
            [cell == "Core" and "Core" or "Other" for cell in self.elem_class[pos]])

XLS_FILE = "../frames-new (1).xlsx"


def test_read_frames():
    fname = XLS_FILE
    fm = FrameMatrix(fname)

    print(fm.frame_types[0])
    pprint(fm.frame_elements[0])
    pprint(fm.elem_class[0])
    print()

def test_frame_matrix():
    fname = XLS_FILE
    fm = FrameMatrix(fname)

    for pos in range(len(fm.frame_types)):
        info = fm.get_frame_info(pos)

        print(info[0])
        print("\t" + "\t".join(info[1]))
        print("\t" + "\t".join(info[2]))
        print("\t" + "\t".join(info[3]))
        print()

def test_frame_info():
    fname = XLS_FILE
    info = FrameInfo(fname)

    print("-" * 20)
    print()

    info.print_frame_info()


def test_access_cells(sheet):

    # access 1 cell
    print(sheet.cell("B2").value)
    print()

    # access w. row/column notation
    #   = cell("C2") where row/col numbers are 0-based
    print(sheet.cell("C2").value)
    print(sheet.cell(row = 2, column = 1).value)
    print()

    # access a cell range
    range = sheet.range("B2:L26")
    for row in range:
        print("\t".join(cell.value for cell in row))

    print()

    # access all rows
    rows = sheet.rows
    for row in rows:
        print(row)

    print()

def main():
    #test_read_frames()

    test_frame_matrix()

    test_frame_info()

if __name__ == "__main__":
    main()
